import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Side, Border, Alignment, PatternFill


def format_data(df, date):
    home_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file = home_path + f"/{date.replace('/', '-')}_DAILY_INCOMING_MFG_WKSHT.xlsx"
    writer = pd.ExcelWriter(file)
    df.to_excel(writer, sheet_name="MFG Totals", engine="openpyxl")
    writer.close()

    wb = load_workbook(filename=file)
    ws = wb.active

    bd = Side(border_style="medium", color="000000")

    def cellSize(cell):
        """letterSize is the width of one letter in excel
            edgeSize is the width of the spaceing at the edge of the cell
            values found using a matrix equation from real values found on excel
            EX: for Calibri at 11 font size with assumtion that equation is
        (excel column width) = letterSize * (number of characters) + edgeSize * (two edges per column)
                [ 2.86 ] = [ 2 2 ] * [ letterSize ]
                [ 4.14 ]   [ 3 2 ] * [ edgeSize   ]
            the matrix was used to solve the two font and edge size variables"""
        value = cell.value
        data_type = cell.data_type
        letterSize = 1.28
        edgeSize = 0.15
        if value is None:
            numChar = 1.0
        elif data_type == "d":
            numChar = 10.0
        elif data_type == "f":
            numChar = 8.0
        else:
            numChar = float(len(str(value)))
        if numChar > 25.0:
            numChar = 25.0
        if numChar < 6.0:
            numChar = 6.0
        cellSize = (letterSize * numChar) + (edgeSize * 2.0)
        return cellSize

    def is_float(s):
        try:
            float(s)
            return True
        except Exception:
            return False

    missing_data_fill = PatternFill(
        start_color="EE1111", end_color="EE1111", fill_type="solid"
    )

    for col in ws.iter_cols(
        min_row=1,
        max_col=(len(df.columns) + 1),
        max_row=(len(df.index) + 1),
    ):
        for cell in col:
            if is_float(cell.value):
                cell.number_format = "$#,##0.00"
            if cell.value is None and col[0].column_letter != "A":
                cell.fill = missing_data_fill
            cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)

    # auto format column widths for both worksheets
    for column_cells in ws.columns:
        length = max((cellSize(cell)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    ws.oddHeader.center.text = f"DAILY INCOMING MFG\nDATE: {date}"
    ws.oddHeader.center.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws["A2"]

    last_column = get_column_letter(len(df.columns) + 1)
    last_row = len(df.index) + 1
    ws.print_area = "A1:" + last_column + str(last_row)
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(file)
